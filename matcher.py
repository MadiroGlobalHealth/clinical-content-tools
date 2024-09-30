"OCL Concepts Matcher to find existing concepts in OCL based on provided Excel metadata"
import json
import math
import time
import openpyxl
from rapidfuzz import process, fuzz
import pandas as pd

# Ignore potential warnings related to opening large Excel files
openpyxl.reader.excel.warnings.simplefilter(action='ignore')

# Load the configuration settings from config.json
with open('config.json', 'r', encoding='utf-8') as f:
    config = json.load(f)

# Extract the configuration settings
# Load the metadata spreadsheet
METADATA_FILEPATH = config.get('METADATA_FILEPATH', './metadata_example.xlsx')
# Load the OCL Concepts spreadsheet
OCL_URL = config.get('OCL_URL', 'https://app.openconceptlab.org/#')
# Matching treshold for fuzzy matching
FUZZY_THRESHOLD = config.get('FUZZY_THRESHOLD', 90)
# Output directory to save the generated form JSONs
OUTPUT_DIR = config.get('OUTPUT_DIR', './generated_form_schemas')
# Get the list of sheets to process from the configuration settings
SHEETS = config.get('SHEETS_TO_MATCH', [])

# Columns names from the metadata spreadsheet
automatch_references = config.get('automatch_references', {})

# Initialize counters
# Counter for total concepts to match
TOTAL_ROWS_PROCESSED = 0

# Initialize a dictionary to store the count of matches found per source
matches_per_source = {}

# Initialize a dictionary to store unique matches per row for all sheets
unique_matches_per_row_all_sheets = {}

# Start the timer to calculate the time taken to run the code
start_time = time.time()

# Find the best matches for each primary and secondary label in the metadata spreadsheet
def find_best_matches(primary, secondary, data, threshold=FUZZY_THRESHOLD, limit=5):
    """
    Find the best matches for a primary and secondary value from a list of data.

    :param primary: The primary value to search for
    :param secondary: The secondary value to search for
    :param data: List of dictionaries with 'id' (string) and 'display_name'
    :param limit: The maximum number of matches to return
    :return: List of tuples containing the id, match, score, and definition
    """
    # Initialize the list of best matches
    results = []

    # Combine the primary and secondary values into a single query string
    query = f"{primary} {secondary}"

    # Extract display_names from data for comparison
    display_names = [item['display_name'] for item in data]

    # Use rapidfuzz's process.extract to find the best matches with threshold
    matches = process.extract(
        query, display_names, scorer=fuzz.WRatio, score_cutoff=threshold, limit=limit
        )

    # Map the matches back to their corresponding IDs and definitions
    for _, score, match_index in matches:
        if score >= FUZZY_THRESHOLD:
            results.append((
                data[match_index]['id'],
                data[match_index]['external_id'],
                data[match_index]['display_name'],
                data[match_index]['description'],
                data[match_index]['datatype'],
                data[match_index]['concept_class'],
                data[match_index]['url'],
                score
                ))
    # Return the list of results
    return results

# Open the metadata Excel file and find the column indices for the required columns
def find_column_index(worksheet, column_name):
    """
    Find the column index for the specified column name in the second row.

    Args:
    worksheet (openpyxl.worksheet.worksheet): The worksheet object.
    column_name (str): The name of the column to find the index for.

    Returns:
    int: The index of the specified column name in the second row.
    """
    for idx, cell in enumerate(worksheet[2], 1):  # assuming the second row contains headers
        if cell.value == column_name:
            return idx
    return -1  # Return -1 if the column name is not found

# Iterate through the sheets in df that are in the sheets list with headers on row 2
for sheet_name in SHEETS:

    df = pd.read_excel(METADATA_FILEPATH, sheet_name=sheet_name, header=1)
    TOTAL_ROWS_PROCESSED += len(df)

    # Iterate through each OCL source and look for suggestions for the primary and secondary lookups
    for source_name, source_config in automatch_references.items():
        print(f"Looking for suggestions in {source_name} source...")
        SOURCE_FILEPATH = source_config['source_filepath']

        with open(SOURCE_FILEPATH, 'r', encoding='UTF-8') as f:
            json_data = json.load(f)
            # Extract only the ID, display names, external IDs, datatype,
            # concept_class, and extras > definitions from the JSON data
            source_data = []
            for item in json_data:
                concept_id = item.get('id', '')
                external_id = item.get('external_id')
                display_name = item.get('display_name', '')
                definition = (
                    item.get('extras', {}).get('definition', '')
                    if item.get('extras', {}) else ''
                )
                description = (
                    item.get('descriptions', [])[0].get('description', '')
                    if item.get('descriptions', []) else definition
                )
                datatype = item.get('datatype')
                concept_class = item.get('concept_class')
                concept_url = item.get('URL')

                # If no description is found, use the definition as the description
                if description == '':
                    description = definition

                # Add the concept details to the source_data list
                source_data.append({
                    'id': concept_id,
                    'display_name': display_name,
                    'definition': definition,
                    'description': description,
                    'datatype': datatype,
                    'concept_class': concept_class,
                    'external_id': external_id,
                    'url': concept_url
                })

        # Load the Excel file, considering the header on row 2
        df = pd.read_excel(METADATA_FILEPATH, sheet_name=sheet_name, header=1)
        print(f"Processing sheet: {sheet_name}")

        # Load the existing Excel file to append the suggestions
        book = openpyxl.load_workbook(METADATA_FILEPATH)

        # Using Excel Writer, append or update the details in the original existing
        # Excel sheet and cells in the specified columns depending on the source name
        with pd.ExcelWriter(METADATA_FILEPATH, engine='openpyxl', mode='a') as writer:
            workbook = writer.book
            ws = workbook[sheet_name]

            # Get the column indices for the suggestion,
            # external ID, description, datatype, concept class, and extras
            suggestion_column = find_column_index(ws, source_config['suggestion_column'])
            external_id_column = find_column_index(ws, source_config['external_id_column'])
            description_column = find_column_index(ws, source_config['description_column'])
            datatype_column = find_column_index(ws, source_config['datatype_column'])
            dataclass_column = find_column_index(ws, source_config['dataclass_column'])
            score_column = find_column_index(ws, source_config['score_column'])

            # Initialize a dictionary to store unique matches per row for the current sheet
            unique_matches_per_row = {}

            # Iterate through each row in the sheet and get
            # the primary and secondary labels to match
            for index, row in df.iterrows():

                primary_lookup = row.get('Label if different') or None
                secondary_lookup = row.get('Question') or row.get('Answers') or None

                # Get suggestions from each OCL source using closest match
                # with RapidFuzz using the primary lookup and secondary lookup
                best_matches = find_best_matches(primary_lookup, secondary_lookup, source_data)
                # If at least one match is found, increment the MATCHES_FOUND counter by 1
                # If at least one match is found, increment the MATCHES_FOUND counter by 1
                if len(best_matches) > 0:
                    # If matches were found, add the unique row identifier to the dictionary
                    # with a value of 1 only if the row identifier
                    # is not already present in the dictionary
                    # Generate unique key using the row index and sheet name
                    # Replace spaces with underscores in the row index and sheet name
                    UNIQUE_ROW_KEY = f"{index}_{sheet_name.replace(' ', '_')}"
                    if UNIQUE_ROW_KEY not in unique_matches_per_row:
                        # Generate unique key using the row index and sheet name
                        unique_matches_per_row[UNIQUE_ROW_KEY] = 1
                    # Update the count of matches found for the current source
                    if source_name in matches_per_source:
                        matches_per_source[source_name] += 1
                    else:
                        matches_per_source[source_name] = 1


                # Add the suggestions to the Excel sheet
                for m in best_matches:
                    # Write the suggestion details to the Excel sheet
                    # Add URL concatenated with OCL_URL in Excel cell using =HYPERLINK() formula
                    ws.cell(row=index+3, column=suggestion_column).value = (
                        f"=HYPERLINK(\"{OCL_URL}{m[6]}\", \"{m[2]}\")"
                    )
                    ws.cell(row=index+3, column=external_id_column).value = m[1]
                    ws.cell(row=index+3, column=description_column).value = m[3]
                    ws.cell(row=index+3, column=datatype_column).value = m[4]
                    ws.cell(row=index+3, column=dataclass_column).value = m[5]
                    ws.cell(row=index+3, column=score_column).value = math.ceil(m[7])

                    print(
                        f"Added suggestion: {m[2]} - {m[1]} with score of {m[7]}"
                    )

            # Close the Excel file writer
            workbook.save(METADATA_FILEPATH)

        # Add the unique_matches_per_row to the unique_matches_per_row_all_sheets dictionary
        unique_matches_per_row_all_sheets.update(unique_matches_per_row)

# Calculate the total number of unique matches found
total_unique_matches_found = sum(unique_matches_per_row_all_sheets.values())

# Calculate the percentage of matches found
percentage_matches_found = (total_unique_matches_found / TOTAL_ROWS_PROCESSED) * 100
rounded_percentage_matches_found = math.ceil(percentage_matches_found)

# Calculate the time taken to run the code
end_time = time.time()
time_taken = end_time - start_time

# Show the final statistics
print("\nFinal Statistics:")
print(f"Total row processed: {TOTAL_ROWS_PROCESSED}")
print(f"Total unique matches found: {total_unique_matches_found}")
print(f"Percentage of matches found: {rounded_percentage_matches_found}%")
print(f"Sources used: {len(automatch_references)}")
print("Matches per source:")
for source_name, matches_per_source in matches_per_source.items():
    print(f"{source_name} source: {matches_per_source} matches")
rounded_time_taken = round(time_taken, 2)
print(f"Time taken to run: {rounded_time_taken} seconds")
